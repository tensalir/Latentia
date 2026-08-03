"""
Embed image previews into a Loop Packaging Creative Intent workbook (v2).

Reads the 'Artwork Folder' configured in Project Info, walks every component
spec tab, and resolves the 'File Name' column against that folder. Supports:
    - Exact match  ("APHRODITE_OUTER.png")
    - Stem match   ("APHRODITE_OUTER" → APHRODITE_OUTER.png/jpg/pdf)
    - Glob match   ("APHRODITE_OUTER_*" or "APHRODITE_OUTER*")
    - Absolute path or path relative to the workbook (back-compat)

Also resolves the 'Packaging Overview Image' file name on Project Info and
the new 'Packing instructions' image rows on each component tab.

Usage
-----
    python refresh_artwork_previews.py <workbook.xlsx>
"""

from __future__ import annotations

import argparse
import shutil
import subprocess
import tempfile
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

NON_SPEC_TABS = {"README", "Project Info", "Components Library", "Product Setup"}

THUMB_MAX_W = 200
THUMB_MAX_H = 90

# Column positions within each component tab's blocks
# These mirror build_template.build_component_tab() exactly.
ARTWORK_HEADER_FILE_COL = 3       # column C
ARTWORK_HEADER_PREVIEW_COL = 5    # column E

PACKING_HEADER_FILE_COL = 3
PACKING_HEADER_PREVIEW_COL = 5

IMAGE_EXTS = (".png", ".jpg", ".jpeg", ".gif", ".bmp", ".tif", ".tiff", ".webp", ".pdf", ".ai")


# ---------------------------------------------------------------------------
def find_in_folder(folder: Path, query: str) -> Path | None:
    """Resolve a file name / partial name / glob against `folder` (recursive).

    Resolution order:
      1. Absolute path that exists.
      2. Relative path that exists from `folder`.
      3. Recursive walk of `folder` matching name / stem / glob.
         Shallower paths win when there are multiple hits, with Archive/ paths
         ranked lowest so a current file is always preferred over a historical
         copy.
    """
    q = (query or "").strip().strip('"').strip("'")
    if not q:
        return None
    p = Path(q)
    if p.is_absolute() and p.exists():
        return p
    if folder and (folder / q).exists():
        return folder / q
    if not folder or not folder.is_dir():
        return None

    def _rank(path: Path) -> tuple[int, int]:
        # Lower score wins. Archive/ paths are penalised; otherwise shallower wins.
        depth = len(path.relative_to(folder).parts)
        archived = any("archive" in part.lower() for part in path.parts)
        return (1 if archived else 0, depth)

    def _is_supplier_output(f: Path) -> bool:
        # Generated supplier PDFs carry the stamped info box; they must never be
        # picked up as Creative Intent artwork previews.
        if any(part.lower() == "supplier_out" for part in f.parts):
            return True
        low = f.name.lower()
        return ("option_a" in low) or ("_overlay" in low) or ("_supplier" in low)

    def _all_image_hits(patterns: list[str]) -> list[Path]:
        hits = []
        for pat in patterns:
            for f in folder.rglob(pat):
                if f.is_file() and f.suffix.lower() in IMAGE_EXTS and not _is_supplier_output(f):
                    hits.append(f)
        return sorted(set(hits), key=_rank)

    # Explicit glob → use as-is.
    if any(ch in q for ch in "*?["):
        hits = _all_image_hits([q])
        return hits[0] if hits else None

    # Exact-name search (any folder).
    hits = _all_image_hits([q])
    if hits:
        return hits[0]
    # Stem + extension search.
    hits = _all_image_hits([f"{q}{ext}" for ext in IMAGE_EXTS])
    if hits:
        return hits[0]
    # Fuzzy prefix.
    hits = _all_image_hits([f"{q}*"])
    return hits[0] if hits else None


def to_thumbnail(src: Path, tmp_dir: Path) -> Path | None:
    suffix = src.suffix.lower()
    target = tmp_dir / (src.stem + ".preview.png")

    if suffix in (".pdf", ".ai"):
        if not shutil.which("pdftoppm"):
            print(f"  ! pdftoppm not available; skipping {src.name}")
            return None
        out_prefix = target.with_suffix("")
        result = subprocess.run(
            ["pdftoppm", "-r", "120", "-png", "-f", "1", "-l", "1",
             "-singlefile", str(src), str(out_prefix)],
            capture_output=True,
        )
        if result.returncode != 0 or not target.exists():
            print(f"  ! pdftoppm failed for {src.name}")
            return None
        img_path = target
    elif suffix in IMAGE_EXTS:
        img_path = src
    else:
        print(f"  ! unsupported file type: {src.name}")
        return None

    try:
        with PILImage.open(img_path) as im:
            im = im.convert("RGB") if im.mode in {"P", "CMYK"} else im
            im.thumbnail((THUMB_MAX_W * 2, THUMB_MAX_H * 2))
            im.save(target, "PNG", optimize=True)
    except Exception as e:
        print(f"  ! cannot open {src}: {e}")
        return None
    return target


def insert_preview(ws, anchor_cell: str, img_path: Path):
    img = XLImage(str(img_path))
    ratio = min(THUMB_MAX_W / img.width, THUMB_MAX_H / img.height, 1.0)
    img.width = int(img.width * ratio)
    img.height = int(img.height * ratio)
    img.anchor = anchor_cell
    ws.add_image(img)


# ---------------------------------------------------------------------------
def _v(cell) -> str:
    return "" if cell.value is None else str(cell.value).strip()


def get_artwork_folder(wb) -> Path | None:
    if "Project Info" not in wb.sheetnames:
        return None
    ws = wb["Project Info"]
    for r in range(5, ws.max_row + 1):
        if _v(ws.cell(row=r, column=2)) == "Artwork Folder":
            val = _v(ws.cell(row=r, column=3))
            if val:
                p = Path(val)
                if p.is_dir():
                    return p
    return None


def find_block_rows(ws, band_keyword: str) -> tuple[int | None, int]:
    """Return (header_row_index, count_of_data_rows) for the named section band.
    Header row is the row directly under the band that holds column labels.
    """
    band_row = None
    for r in range(1, ws.max_row + 1):
        v = _v(ws.cell(row=r, column=1))
        if v.startswith(band_keyword):
            band_row = r
            break
    if band_row is None:
        return None, 0
    header_row = band_row + 1
    # Count data rows until the next gray-banded section
    count = 0
    r = header_row + 1
    while r <= ws.max_row:
        # Stop if we run into the next section band (column A value is a band-like text)
        v_a = _v(ws.cell(row=r, column=1))
        # The next band would be 'Packing instructions' or 'Dimensions'
        if v_a.startswith("Packing instructions") or v_a.startswith("Dimensions") or v_a.startswith("Artwork files"):
            break
        # Stop if we hit a totally empty row beyond a sane window
        if not v_a and not _v(ws.cell(row=r, column=2)) and not _v(ws.cell(row=r, column=3)):
            if count > 0:
                # allow one blank row but stop on second
                count_blank = 0
                rr = r
                while rr <= ws.max_row:
                    if not any(_v(ws.cell(row=rr, column=col)) for col in range(1, 6)):
                        count_blank += 1
                        if count_blank > 1:
                            break
                    else:
                        break
                    rr += 1
                break
        count += 1
        r += 1
    return header_row, count


# ---------------------------------------------------------------------------
def refresh(workbook_path: Path) -> None:
    workbook_path = workbook_path.resolve()
    workbook_dir = workbook_path.parent
    print(f"Refreshing previews in: {workbook_path}")

    wb = load_workbook(workbook_path)
    folder = get_artwork_folder(wb) or workbook_dir
    print(f"Artwork folder: {folder}")

    tmp_dir = Path(tempfile.mkdtemp(prefix="loop_artwork_"))
    inserted = 0
    skipped = 0
    missing = []

    # Project Info — packaging overview image
    if "Project Info" in wb.sheetnames:
        ws = wb["Project Info"]
        for r in range(5, ws.max_row + 1):
            if _v(ws.cell(row=r, column=2)) == "Packaging Overview Image":
                fname = _v(ws.cell(row=r, column=3))
                if fname:
                    resolved = find_in_folder(folder, fname)
                    if resolved is None:
                        # try workbook-relative as last resort
                        p = (workbook_dir / fname)
                        if p.exists():
                            resolved = p
                    if resolved:
                        thumb = to_thumbnail(resolved, tmp_dir)
                        if thumb:
                            ws._images = [im for im in (ws._images or [])
                                          if getattr(im, "anchor", None) != f"D{r}"]
                            insert_preview(ws, f"D{r}", thumb)
                            inserted += 1
                    else:
                        missing.append(("Project Info", fname))
                break

    # Component tabs
    for sheet_name in wb.sheetnames:
        if sheet_name in NON_SPEC_TABS:
            continue
        ws = wb[sheet_name]
        ws._images = []  # reset and rebuild

        # Artwork files block
        art_header, art_count = find_block_rows(ws, "Artwork files")
        if art_header:
            for i in range(1, art_count + 1):
                rr = art_header + i
                fname = _v(ws.cell(row=rr, column=ARTWORK_HEADER_FILE_COL))
                if not fname:
                    continue
                resolved = find_in_folder(folder, fname)
                if not resolved:
                    missing.append((sheet_name, fname)); skipped += 1; continue
                thumb = to_thumbnail(resolved, tmp_dir)
                if not thumb:
                    skipped += 1; continue
                insert_preview(ws,
                               f"{get_column_letter(ARTWORK_HEADER_PREVIEW_COL)}{rr}",
                               thumb)
                inserted += 1

        # Packing instructions block (optional per component)
        pi_header, pi_count = find_block_rows(ws, "Packing instructions")
        if pi_header:
            for i in range(1, pi_count + 1):
                rr = pi_header + i
                fname = _v(ws.cell(row=rr, column=PACKING_HEADER_FILE_COL))
                if not fname:
                    continue
                resolved = find_in_folder(folder, fname)
                if not resolved:
                    missing.append((sheet_name, fname)); skipped += 1; continue
                thumb = to_thumbnail(resolved, tmp_dir)
                if not thumb:
                    skipped += 1; continue
                insert_preview(ws,
                               f"{get_column_letter(PACKING_HEADER_PREVIEW_COL)}{rr}",
                               thumb)
                inserted += 1

    wb.save(workbook_path)
    print(f"\nEmbedded: {inserted}  Skipped: {skipped}  Missing: {len(missing)}")
    for tab, name in missing[:20]:
        print(f"  [missing] {tab}: {name}")


def main():
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("workbook", type=Path)
    refresh(p.parse_args().workbook)


if __name__ == "__main__":
    main()
