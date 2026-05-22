"""
Loop Packaging Creative Intent — Excel template builder (v2).

Generates the empty template. The Nyx-prefilled example is built by
build_nyx_example.py which calls into this module.

v2 changes (per Ana's feedback)
  - Project Info now includes 'Artwork Folder' and 'Packaging Overview Image'.
  - Single SKU/Colourway metadata field; the multi-row variants table is gone
    because one workbook = one SKU.
  - Component tabs collapse from up-to-8 option columns to a single spec column.
  - Artwork section uses pre-named slots (Mockup, Artwork_Front, Artwork_Back,
    Artwork) so the supplier knows exactly what to look for; 'File Path' is now
    'File Name' (a name to search for inside the Artwork Folder).
  - New 'Packing Instructions' block on every component tab: rows of
    Step / Instruction / Image File Name / Preview.

Tab layout (per component)
    BLOCK A  Header (display name, description, PDF page title)
    BLOCK B  Specifications (one column)
    BLOCK C  Artwork files (fixed slots + optional extras)
    BLOCK D  Packing instructions (free-form steps with image references)
    BLOCK E  Dimensions
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT_DIR = Path(".")
TEMPLATE_PATH = OUT_DIR / "Loop_Packaging_Creative_Intent_TEMPLATE.xlsx"

# -- Style tokens -----------------------------------------------------------
FONT_NAME = "Arial"
INK_BLACK = "1A1A1A"
INK_DARK = "333333"
INK_MID = "6B6B6B"
LINE_GRAY = "D0D0D0"
BAND_GRAY = "F2F2F2"
BAND_DARK = "262626"
ACCENT = "C8102E"

THIN = Side(style="thin", color=LINE_GRAY)
THIN_DARK = Side(style="thin", color=INK_DARK)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def f(size=10, bold=False, color=INK_BLACK, italic=False):
    return Font(name=FONT_NAME, size=size, bold=bold, color=color, italic=italic)


def fill(rgb):
    return PatternFill("solid", start_color=rgb, end_color=rgb)


LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)


# -- Reference data ---------------------------------------------------------
SPEC_FIELDS = [
    "Drawing Part Number",
    "Print Part Number",
    "Material",
    "Inks / Print",
    "Finishes",
    "Special Effects",
    "Printing Method",
    "Coating MSDS Ref.",
    "Approval Status",
    "Notes",
]

PRINTING_METHODS = ["Offset", "Flexo", "Digital", "Screen", "N/A"]
APPROVAL_STATES = ["Draft", "In review", "Approved", "Blocked"]

# Default artwork slot configuration per "component style".
# Style A — printed sheet with two visible faces (Outer Sleeve, Inner Tray).
# Style B — single-face item with packing context (Tissue, Sticker, Closure).
ARTWORK_SLOTS = {
    "two_face": ["Mockup", "Artwork_Front", "Artwork_Back"],
    "single_face": ["Mockup", "Artwork"],
}

# Component definitions: (tab_name, display, description, slots-style, has_packing_block)
NYX_COMPONENTS = [
    ("Outer_Sleeve",   "Outer Sleeve",   "Outer paper sleeve that wraps the inner tray.",                  "two_face",    False),
    ("Inner_Tray",     "Inner Tray",     "Folded paper tray that holds the eye mask + accessories.",       "two_face",    False),
    ("Tissue_Paper",   "Tissue Paper",   "Pre-dyed inner wrap securing the product inside the tray.",      "single_face", True),
    ("Tissue_Sticker", "Tissue Sticker", "Decorative or content sticker that closes the tissue wrap.",     "single_face", True),
    ("Closure_Sticker","Closure Sticker","Tear-strip sticker with EAN barcode and SKU colourway info.",    "single_face", True),
]

PORTFOLIO_COMPONENTS = [
    ("Hangtag",          "Hangtag",          "Retail/D2C hangtag — rename or reuse this slot.",  "single_face", False),
    ("Insert_Card",      "Insert Card",      "Information / thank-you card.",                    "two_face",    False),
    ("Earplug_Case",     "Earplug Case",     "Hard or soft case for earplug SKUs.",              "two_face",    False),
    ("Carry_Pouch",      "Carry Pouch",      "Soft carry pouch.",                                "single_face", False),
    ("Polybag",          "Polybag",          "Outer protective polybag if applicable.",          "single_face", False),
    ("Component_Slot_A", "Component Slot A", "Empty slot — rename for any future component.",    "two_face",    True),
    ("Component_Slot_B", "Component Slot B", "Empty slot — rename for any future component.",    "single_face", True),
]

ALL_COMPONENTS = NYX_COMPONENTS + PORTFOLIO_COMPONENTS


# -- Workbook helpers -------------------------------------------------------
def new_workbook() -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def page_setup(ws):
    ws.sheet_view.showGridLines = False
    ws.print_options.horizontalCentered = True
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4


def title_band(ws, text, row, col_start=1, col_end=10, height=28):
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=col_start, value=text)
    c.font = f(size=14, bold=True, color="FFFFFF")
    c.fill = fill(BAND_DARK)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = height


def section_band(ws, text, row, col_start=1, col_end=10, height=22):
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=col_start, value=text)
    c.font = f(size=11, bold=True, color=INK_BLACK)
    c.fill = fill(BAND_GRAY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border = Border(bottom=THIN_DARK)
    ws.row_dimensions[row].height = height


def label_cell(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.font = f(size=10, bold=True)
    c.alignment = LEFT
    c.fill = fill(BAND_GRAY)
    c.border = BOX
    return c


def value_cell(ws, row, col, value=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = f(size=10)
    c.alignment = LEFT_TOP
    c.border = BOX
    return c


# -- Tabs -------------------------------------------------------------------
def build_readme(wb):
    ws = wb.create_sheet("README")
    page_setup(ws)
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 110

    title_band(ws, "Loop Packaging Creative Intent — Template (v2)", 1, col_end=2, height=34)
    r = 3
    sections = [
        ("What this workbook is",
         "Reusable template for Loop's Packaging Creative Intent documents. "
         "One workbook = one SKU. Fill it in, then run the generator to produce "
         "the supplier-ready PDF in the Nyx layout."),
        ("Set-up — do this once per product",
         "1. Project Info — fill the metadata block. The two important new fields are:\n"
         "   • Artwork Folder — the folder on disk that holds all your artwork files for this SKU\n"
         "   • Packaging Overview Image — file name (no path) of the exploded/overview render that lives in that folder\n"
         "2. SKU / Colourway — single value (e.g. 'Black' or 'Blue'). For a second colourway, duplicate the workbook.\n"
         "3. Components Library — review the master list; add rows for any new component types your portfolio uses.\n"
         "4. Product Setup — mark which components apply to THIS product, set page order.\n"
         "5. Component tabs — for each, fill the spec block and the artwork file-name slots."),
        ("Artwork file names — how the search works",
         "On every artwork row you only need to type the FILE NAME (or a partial name). "
         "The refresh script and the PDF generator both look inside the 'Artwork Folder' you set on Project Info and "
         "match by name + glob fallback (e.g. 'APHRODITE_OUTER' matches APHRODITE_OUTER.png, APHRODITE_OUTER_v3.pdf, etc.). "
         "Absolute or relative paths still work if you prefer."),
        ("Two component styles",
         "Style A (two-face) — Outer Sleeve, Inner Tray, Insert Card, Earplug Case. Renders Mockup + Artwork_Front + Artwork_Back. "
         "Leave Artwork_Back blank if not applicable; it will be skipped automatically.\n"
         "Style B (single-face) — Tissue, Sticker, Closure Sticker, Polybag, etc. Renders Mockup + Artwork + Packing Instructions."),
        ("Embedding artwork previews",
         "After you've named your files, run:\n"
         "    python refresh_artwork_previews.py <this_workbook.xlsx>\n"
         "It walks the workbook, finds each file in the Artwork Folder, and embeds a thumbnail in the Preview cell. "
         "Re-run any time you swap a file."),
        ("Generating the PDF",
         "When the workbook is filled and previews are embedded:\n"
         "    python generate_creative_intent_pdf.py <this_workbook.xlsx>\n"
         "Output is named after the project (suffix '_OUTPUT.pdf'). One PDF per SKU."),
    ]
    for title, body in sections:
        c = ws.cell(row=r, column=2, value=title)
        c.font = f(size=12, bold=True, color=INK_BLACK)
        c.alignment = LEFT
        ws.row_dimensions[r].height = 20
        r += 1
        c = ws.cell(row=r, column=2, value=body)
        c.font = f(size=10, color=INK_DARK)
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[r].height = max(40, body.count("\n") * 16 + 30)
        r += 2


def build_project_info(wb):
    ws = wb.create_sheet("Project Info")
    page_setup(ws)
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 70
    ws.column_dimensions["D"].width = 55

    title_band(ws, "Project Info", 1, col_end=4, height=32)

    fields = [
        ("Project Name", "e.g. Nyx Packaging"),
        ("Product Type", "e.g. Sleep Mask, Earplug, Accessory"),
        ("Product Family", "e.g. Sleep, Focus, Music"),
        ("SKU / Colourway", "Single value, e.g. 'Black' or 'Blue'. One workbook = one SKU."),
        ("Packaging Designer", "Person responsible for the packaging artwork"),
        ("Packaging Engineer", "Person responsible for the structural design"),
        ("Brand Manager", "Optional — owner of the launch"),
        ("Date", "DD/MM/YYYY"),
        ("Project Stage", "EVT / DVT / PVT / MP"),
        ("Supplier", "Print supplier name + contact"),
        ("Internal Reference", "Project tracker ID (Asana, Monday, etc.)"),
        ("Artwork Folder", "Absolute path to the folder that holds artwork files for this SKU"),
        ("Packaging Overview Image", "File name of the exploded/overview render (e.g. APHRODITE_overview.png)"),
        ("Notes", "Anything the supplier or downstream team should know"),
    ]

    section_band(ws, "Top-level metadata", 3, col_end=4)
    label_cell(ws, 4, 2, "Field")
    label_cell(ws, 4, 3, "Value")
    label_cell(ws, 4, 4, "Hint")
    r = 5
    for label, hint in fields:
        label_cell(ws, r, 2, label)
        value_cell(ws, r, 3, "")
        h = value_cell(ws, r, 4, hint)
        h.font = f(size=9, italic=True, color=INK_MID)
        ws.row_dimensions[r].height = 22
        r += 1


def build_components_library(wb):
    ws = wb.create_sheet("Components Library")
    page_setup(ws)
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 26
    ws.column_dimensions["E"].width = 55
    ws.column_dimensions["F"].width = 18

    title_band(ws, "Components Library", 1, col_end=6, height=32)
    section_band(ws,
                 "Master list of every packaging component used across Loop's portfolio. "
                 "Add a new row any time you introduce a new component type.",
                 3, col_end=6)
    r = 4
    for i, h in enumerate(["Component ID", "Tab Name", "Display Name", "Description", "Style"]):
        label_cell(ws, r, 2 + i, h)
    r = 5
    for idx, (tab_name, display, desc, style, _packing) in enumerate(ALL_COMPONENTS, start=1):
        comp_id = f"C{idx:03d}"
        for col, val in enumerate([comp_id, tab_name, display, desc, style]):
            v = value_cell(ws, r, 2 + col, val)
            if col == 0:
                v.font = f(size=10, bold=True)
        ws.row_dimensions[r].height = 22
        r += 1
    for _ in range(5):
        for col in range(2, 7):
            value_cell(ws, r, col, "")
        ws.row_dimensions[r].height = 20
        r += 1


def build_product_setup(wb):
    ws = wb.create_sheet("Product Setup")
    page_setup(ws)
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 26
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 50

    title_band(ws, "Product Setup — which components apply to THIS product?", 1, col_end=7, height=32)
    section_band(ws,
                 "Tick 'Yes' for every component this product uses. "
                 "Page Order controls the sequence in the generated PDF.",
                 3, col_end=7)
    r = 4
    for i, h in enumerate(["Component ID", "Tab Name", "Display Name", "Include?", "Page Order", "Per-product notes"]):
        label_cell(ws, r, 2 + i, h)
    r = 5
    yn_dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(yn_dv)
    for idx, (tab_name, display, desc, style, _packing) in enumerate(ALL_COMPONENTS, start=1):
        comp_id = f"C{idx:03d}"
        is_nyx = (tab_name, display, desc, style, _packing) in NYX_COMPONENTS
        for col, val in enumerate([comp_id, tab_name, display]):
            v = value_cell(ws, r, 2 + col, val)
            if col == 0:
                v.font = f(size=10, bold=True)
        v = value_cell(ws, r, 5, "Yes" if is_nyx else "No")
        yn_dv.add(v)
        order = (idx if is_nyx else "")
        value_cell(ws, r, 6, order)
        value_cell(ws, r, 7, "")
        ws.row_dimensions[r].height = 22
        r += 1


def build_component_tab(wb, tab_name, display, description, style, has_packing):
    """One tab per component (v2 — single-option spec, named artwork slots)."""
    ws = wb.create_sheet(tab_name[:31])
    page_setup(ws)
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 26
    ws.column_dimensions["D"].width = 50
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 22
    ws.column_dimensions["G"].width = 22

    # BLOCK A — header
    title_band(ws, f"{display} — Component Spec", 1, col_end=7, height=32)
    section_band(ws, "Component header", 3, col_end=7)
    label_cell(ws, 4, 1, "Display Name");      value_cell(ws, 4, 2, display)
    ws.merge_cells(start_row=4, start_column=2, end_row=4, end_column=7)
    label_cell(ws, 5, 1, "Description");       value_cell(ws, 5, 2, description)
    ws.merge_cells(start_row=5, start_column=2, end_row=5, end_column=7)
    label_cell(ws, 6, 1, "PDF Page Title");    value_cell(ws, 6, 2, "")
    ws.merge_cells(start_row=6, start_column=2, end_row=6, end_column=7)
    for r in (4, 5, 6):
        ws.row_dimensions[r].height = 24

    # BLOCK B — specifications (single column)
    section_band(ws, "Specifications", 8, col_end=7)
    label_cell(ws, 9, 1, "Field")
    label_cell(ws, 9, 2, "Value")
    ws.row_dimensions[9].height = 22

    method_dv = DataValidation(type="list", formula1='"' + ",".join(PRINTING_METHODS) + '"', allow_blank=True)
    status_dv = DataValidation(type="list", formula1='"' + ",".join(APPROVAL_STATES) + '"', allow_blank=True)
    ws.add_data_validation(method_dv)
    ws.add_data_validation(status_dv)

    for fi, field in enumerate(SPEC_FIELDS):
        row = 10 + fi
        label_cell(ws, row, 1, field)
        v = value_cell(ws, row, 2, "")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
        if field == "Printing Method":
            method_dv.add(v)
        elif field == "Approval Status":
            status_dv.add(v)
        ws.row_dimensions[row].height = 28

    # BLOCK C — artwork slots
    art_start = 10 + len(SPEC_FIELDS) + 1
    section_band(ws, "Artwork files (file name OR full path; refresh script embeds the preview)", art_start, col_end=7)
    header_row = art_start + 1
    for col, h in enumerate(["Artwork Type", "Caption", "File Name", "Preview"]):
        if col == 0:
            label_cell(ws, header_row, 1, h)
        elif col == 1:
            label_cell(ws, header_row, 2, h)
        elif col == 2:
            label_cell(ws, header_row, 3, h)
            ws.merge_cells(start_row=header_row, start_column=3, end_row=header_row, end_column=4)
        else:
            label_cell(ws, header_row, 5, h)
            ws.merge_cells(start_row=header_row, start_column=5, end_row=header_row, end_column=7)
    ws.row_dimensions[header_row].height = 22

    slot_types = ARTWORK_SLOTS[style]
    # Render the fixed slots + 3 extra "free" rows for special items
    artwork_rows = list(slot_types) + ["", "", ""]
    for i, atype in enumerate(artwork_rows):
        rr = header_row + 1 + i
        value_cell(ws, rr, 1, atype)
        ws.cell(row=rr, column=1).font = f(size=10, bold=bool(atype))
        value_cell(ws, rr, 2, "")  # caption
        value_cell(ws, rr, 3, "")  # file name
        ws.merge_cells(start_row=rr, start_column=3, end_row=rr, end_column=4)
        value_cell(ws, rr, 5, "")  # preview cell
        ws.merge_cells(start_row=rr, start_column=5, end_row=rr, end_column=7)
        ws.row_dimensions[rr].height = 90

    last_art_row = header_row + len(artwork_rows)

    # BLOCK D — packing instructions (only on relevant styles)
    next_row = last_art_row + 2
    if has_packing or style == "single_face":
        section_band(ws, "Packing instructions (text + reference image)", next_row, col_end=7)
        pi_header = next_row + 1
        for col, h in enumerate(["Step", "Instruction", "Image File Name", "Preview"]):
            if col == 0:
                label_cell(ws, pi_header, 1, h)
            elif col == 1:
                label_cell(ws, pi_header, 2, h)
            elif col == 2:
                label_cell(ws, pi_header, 3, h)
                ws.merge_cells(start_row=pi_header, start_column=3, end_row=pi_header, end_column=4)
            else:
                label_cell(ws, pi_header, 5, h)
                ws.merge_cells(start_row=pi_header, start_column=5, end_row=pi_header, end_column=7)
        ws.row_dimensions[pi_header].height = 22
        for i in range(5):
            rr = pi_header + 1 + i
            value_cell(ws, rr, 1, f"Step {i+1}")
            ws.cell(row=rr, column=1).font = f(size=10, bold=True)
            value_cell(ws, rr, 2, "")
            value_cell(ws, rr, 3, "")
            ws.merge_cells(start_row=rr, start_column=3, end_row=rr, end_column=4)
            value_cell(ws, rr, 5, "")
            ws.merge_cells(start_row=rr, start_column=5, end_row=rr, end_column=7)
            ws.row_dimensions[rr].height = 80
        next_row = pi_header + 1 + 5 + 1
    else:
        next_row = last_art_row + 2

    # BLOCK E — dimensions
    section_band(ws, "Dimensions", next_row, col_end=7)
    dim_rows = ["Height (mm)", "Width (mm)", "Depth (mm)", "Net weight (g)", "Sticker / element placement"]
    rr = next_row + 1
    for label in dim_rows:
        label_cell(ws, rr, 1, label)
        value_cell(ws, rr, 2, "")
        ws.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=7)
        ws.row_dimensions[rr].height = 24 if "placement" not in label.lower() else 40
        rr += 1

    return ws


# -- Build entry point ------------------------------------------------------
def build(path: Path):
    wb = new_workbook()
    build_readme(wb)
    build_project_info(wb)
    build_components_library(wb)
    build_product_setup(wb)
    for tab_name, display, desc, style, has_packing in ALL_COMPONENTS:
        build_component_tab(wb, tab_name, display, desc, style, has_packing)
    wb.save(path)


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("output", type=Path, nargs="?", default=TEMPLATE_PATH,
                        help="Output path for the empty template .xlsx")
    args = parser.parse_args()
    build(args.output)
    print(f"Saved: {args.output}")
