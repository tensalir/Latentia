#!/usr/bin/env python3
"""
Sync a Creative Intent workbook with its stage folder and .ai artwork files.

This is the "make the workbook match reality" step that sits between filling
the workbook by hand and generating the PDFs. One command does all of it:

  1. Sets Project Info -> Artwork Folder to the stage root (the one path
     every other script resolves file names against).
  2. Converts any .psd renders under the stage root to .png so the preview
     embedder can use them (it can't read Photoshop files).
  3. Rebuilds any component tab that is Included in Product Setup but missing
     from the workbook (tabs get dropped by accident during Google Sheets
     round-trips). Display name / description / single_face vs two_face come
     from the Components Library; default artwork file-name patterns are
     pre-filled ({Tab}_Mockup and {Tab}*_ED*).
  4. Removes the retired "Special Effects" spec row from every component tab
     (older workbooks still carry it).
  5. For each --ai-file (or every .ai found under the stage's Print_Files/):
     matches it to a component tab by longest tab-name prefix, then writes
     the Illustrator -> Excel fields:
        Inks / Print        <- ink plate names
        Finishes            <- special-finish plate names
        Print Part Number   <- the .ai file stem
        Notes               <- "Structural plates: ..." (dieline plate names)
  6. Reorders component tabs to match Product Setup page order.

Components included in Product Setup that have no .ai yet are left alone and
reported — they render as [no artwork] in the Creative Intent until their
files arrive, which is the intended behaviour for planned-but-not-ready parts.

Usage
-----
python sync_workbook.py <workbook.xlsx> --stage-root <path/to/EVT> [--ai-file X.ai ...]

If --ai-file is omitted, every *.ai under <stage-root>/Print_Files/ is used.
"""
from __future__ import annotations

import argparse
import importlib.util
import sys
from pathlib import Path

import openpyxl

HERE = Path(__file__).resolve().parent
META_TABS = {"README", "Project Info", "Components Library", "Product Setup"}


def _load(module_filename: str):
    path = HERE / module_filename
    spec = importlib.util.spec_from_file_location(path.stem, path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def component_tabs(wb) -> list[str]:
    return [s for s in wb.sheetnames if s not in META_TABS]


def match_component_tab(ai_stem: str, tabs: list[str]) -> str | None:
    """Longest tab-name prefix wins. Double underscores (the variant
    separator, e.g. Rigid_Box_Lid__Black) collapse for matching."""
    norm = ai_stem.lower().replace("__", "_")
    best = None
    for tab in tabs:
        if norm.startswith(tab.lower()):
            if best is None or len(tab) > len(best):
                best = tab
    return best


def _set_project_info(ws, label: str, value: str) -> bool:
    for r in range(1, 30):
        if ws.cell(row=r, column=2).value == label:
            ws.cell(row=r, column=3).value = value
            return True
    return False


def _set_spec(ws, label: str, value: str) -> bool:
    for r in range(1, 40):
        if ws.cell(row=r, column=1).value == label:
            ws.cell(row=r, column=2).value = value
            return True
    return False


def convert_psds(folder: Path) -> list[str]:
    done = []
    try:
        from PIL import Image
    except ImportError:
        return done
    for psd in folder.rglob("*.psd"):
        png = psd.with_suffix(".png")
        if png.exists():
            continue
        try:
            Image.open(psd).convert("RGB").save(png)
            done.append(png.name)
        except Exception:
            pass
    return done


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("workbook", type=Path)
    ap.add_argument("--stage-root", type=Path, required=True,
                    help="The stage folder (e.g. .../02_Production/EVT)")
    ap.add_argument("--ai-file", type=Path, action="append", default=[],
                    help="Repeatable. Defaults to every .ai under Print_Files/")
    args = ap.parse_args()

    stage = args.stage_root.resolve()
    wb_path = args.workbook.resolve()

    bt = _load("build_template.py")
    sup = _load("generate_supplier_pdf.py")

    # 1 + 2 — artwork folder and PSD conversion
    converted = convert_psds(stage)
    for name in converted:
        print(f"converted PSD -> {name}")

    wb = openpyxl.load_workbook(wb_path)
    if "Project Info" in wb.sheetnames:
        _set_project_info(wb["Project Info"], "Artwork Folder", str(stage))

    # 3 — rebuild missing included tabs
    ps = wb["Product Setup"]
    included = []  # (order, tab, display)
    for r in range(5, ps.max_row + 1):
        tab = ps.cell(row=r, column=3).value
        disp = ps.cell(row=r, column=4).value
        inc = ps.cell(row=r, column=5).value
        po = ps.cell(row=r, column=6).value
        if tab and str(inc).strip().lower() == "yes":
            included.append((float(po) if po is not None else 999, tab, disp))
    included.sort()

    lib = {}
    wl = wb["Components Library"]
    for r in range(5, wl.max_row + 1):
        tab = wl.cell(row=r, column=3).value
        if tab:
            lib[tab] = (wl.cell(row=r, column=4).value,
                        wl.cell(row=r, column=5).value,
                        wl.cell(row=r, column=6).value)

    for _order, tab, disp in included:
        if tab in wb.sheetnames:
            continue
        display, desc, style = lib.get(tab, (disp, "", "single_face"))
        style = style if style in ("single_face", "two_face") else "single_face"
        ws_new = bt.build_component_tab(wb, tab[:31], display, desc or "",
                                        style, style == "single_face")
        # pre-fill default artwork file-name patterns
        for r in range(1, ws_new.max_row + 1):
            if ws_new.cell(row=r, column=1).value == "Artwork Type":
                for i in range(1, 6):
                    ttype = ws_new.cell(row=r + i, column=1).value
                    if ttype == "Mockup":
                        ws_new.cell(row=r + i, column=3).value = f"{tab}_Mockup"
                    elif ttype in ("Artwork", "Artwork_Front"):
                        ws_new.cell(row=r + i, column=3).value = f"{tab}*_ED*"
                break
        print(f"rebuilt missing tab: {tab} ({style})")

    # 4 — retire Special Effects everywhere
    for tab in component_tabs(wb):
        ws = wb[tab]
        for r in range(1, 40):
            if ws.cell(row=r, column=1).value == "Special Effects":
                ws.delete_rows(r, 1)
                break

    # 6 — tab order = Product Setup page order
    order_names = ["README", "Project Info", "Components Library", "Product Setup"]
    order_names += [t for _, t, _ in included if t in wb.sheetnames]
    order_names += [t for t in wb.sheetnames if t not in order_names]
    name_to_ws = {w.title: w for w in wb.worksheets}
    wb._sheets = [name_to_ws[n] for n in order_names if n in name_to_ws]
    wb.save(wb_path)

    # 5 — Illustrator -> Excel per AI file
    ai_files = [p.resolve() for p in args.ai_file]
    if not ai_files:
        ai_files = sorted((stage / "Print_Files").rglob("*.ai"))
    tabs = component_tabs(openpyxl.load_workbook(wb_path))
    matched = set()
    for ai in ai_files:
        tab = match_component_tab(ai.stem, tabs)
        if not tab:
            print(f"  ! no tab match for {ai.name} — skipped")
            continue
        plate = sup.extract_plate_info(ai)
        wb2 = openpyxl.load_workbook(wb_path)
        ws = wb2[tab]
        _set_spec(ws, "Inks / Print", ", ".join(plate.get("inks", [])))
        _set_spec(ws, "Finishes", ", ".join(plate.get("finishes", [])))
        _set_spec(ws, "Print Part Number", ai.stem)
        if plate.get("dielines"):
            _set_spec(ws, "Notes",
                      "Structural plates: " + ", ".join(plate["dielines"]))
        wb2.save(wb_path)
        matched.add(tab)
        print(f"  {tab:28} inks={len(plate.get('inks', []))} "
              f"finishes={len(plate.get('finishes', []))} "
              f"dielines={len(plate.get('dielines', []))}")

    pending = [t for _, t, _ in included if t not in matched]
    if pending:
        print(f"no artwork yet (will render as [no artwork]): {', '.join(pending)}")
    print(f"Synced: {wb_path}")


if __name__ == "__main__":
    main()
