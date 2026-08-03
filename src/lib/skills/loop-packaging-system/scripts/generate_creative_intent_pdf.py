"""
Generate the Packaging Creative Intent PDF from a filled Excel workbook (v2).

Per-page layout
  Page 1 (Overview): the Packaging Overview Image (exploded render) placed
    large on the right, with the component list down the left as a key.
  Pages 2+ (Components): single-option spec block at the top, then the
    relevant artwork(s) with their file names + an optional Packing
    Instructions section.

Component styles (driven by Components Library / Product Setup)
  two_face     — Outer Sleeve / Inner Tray etc.  → Mockup + Artwork_Front + Artwork_Back
  single_face  — Tissue / Sticker / Closure etc. → Mockup + Artwork + Packing Instructions

Usage
-----
    python generate_creative_intent_pdf.py <workbook.xlsx> [output.pdf]
"""

from __future__ import annotations

import argparse
import shutil
import subprocess
import tempfile
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image as PILImage

from reportlab.lib.colors import HexColor, white
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas

# -- Page geometry ----------------------------------------------------------
PAGE_W, PAGE_H = landscape(A4)  # 842 x 595 pt
MARGIN = 16 * mm

INK = HexColor("#1A1A1A")
INK_DARK = HexColor("#262626")
INK_MID = HexColor("#6B6B6B")
LINE = HexColor("#D0D0D0")
BAND = HexColor("#F2F2F2")
ACCENT = HexColor("#C8102E")

FONT_REG = "Helvetica"
FONT_BOLD = "Helvetica-Bold"
FONT_ITALIC = "Helvetica-Oblique"
FONT_MONO = "Courier"

# Spec lines rendered in the PDF (label as it appears on the page).
RENDERED_SPEC_LINES = [
    ("Drawing Part Number", "DRAWING PART NUMBER"),
    ("Print Part Number",   "PRINT PART NUMBER"),
    ("Material",            "MATERIAL"),
    ("Inks / Print",        "INKS"),
    ("Finishes",            "FINISHES"),
    ("Printing Method",     "PRINTING METHOD"),
]

IMAGE_EXTS = (".png", ".jpg", ".jpeg", ".gif", ".bmp", ".tif", ".tiff", ".webp", ".pdf")


# -- Workbook readers -------------------------------------------------------
@dataclass
class ArtworkEntry:
    artwork_type: str
    caption: str
    file_name: str


@dataclass
class PackingStep:
    step: str
    instruction: str
    file_name: str


@dataclass
class Component:
    tab_name: str
    display_name: str
    description: str
    page_order: int
    style: str = "single_face"   # 'two_face' or 'single_face'
    pdf_page_title: str = ""
    specs: dict = field(default_factory=dict)            # field -> value (str)
    artworks: list[ArtworkEntry] = field(default_factory=list)
    packing_steps: list[PackingStep] = field(default_factory=list)
    dimensions: dict = field(default_factory=dict)


@dataclass
class Project:
    project_name: str = ""
    product_type: str = ""
    product_family: str = ""
    sku_colourway: str = ""
    designer: str = ""
    engineer: str = ""
    graphic_designer: str = ""
    brand_manager: str = ""
    date: str = ""
    stage: str = ""
    supplier: str = ""
    internal_ref: str = ""
    notes: str = ""
    artwork_folder: Path | None = None
    overview_image_name: str = ""
    components: list[Component] = field(default_factory=list)


def _v(cell) -> str:
    return "" if cell.value is None else str(cell.value).strip()


def _format_date_eu(raw) -> str:
    """DD-MM-YYYY, no time. Accepts datetime/date/string; falls back gracefully."""
    import datetime as _dt
    if raw is None or raw == "":
        return ""
    if isinstance(raw, (_dt.datetime, _dt.date)):
        return raw.strftime("%d-%m-%Y")
    s = str(raw).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y",
                "%m/%d/%Y", "%Y/%m/%d", "%d.%m.%Y"):
        try:
            return _dt.datetime.strptime(s, fmt).strftime("%d-%m-%Y")
        except ValueError:
            continue
    return s.split(" ")[0]


def read_project_info(ws):
    info = {}
    for r in range(5, ws.max_row + 1):
        label = _v(ws.cell(row=r, column=2))
        if not label:
            continue
        info[label] = _v(ws.cell(row=r, column=3))
    return info


def read_components_library(ws) -> dict:
    """Return tab_name → style mapping."""
    styles = {}
    for r in range(5, ws.max_row + 1):
        tab = _v(ws.cell(row=r, column=3))
        style = _v(ws.cell(row=r, column=6))
        if tab and style:
            styles[tab] = style
    return styles


def read_product_setup(ws) -> list[dict]:
    rows = []
    for r in range(5, ws.max_row + 1):
        tab = _v(ws.cell(row=r, column=3))
        display = _v(ws.cell(row=r, column=4))
        include = _v(ws.cell(row=r, column=5)).lower() == "yes"
        order_raw = _v(ws.cell(row=r, column=6))
        try:
            order = int(order_raw) if order_raw else 9999
        except ValueError:
            order = 9999
        if tab:
            rows.append({"tab_name": tab, "display": display,
                         "include": include, "order": order})
    return rows


def _find_block(ws, band_keyword: str) -> tuple[int | None, int]:
    """Locate a section band and return (header_row, data_row_count)."""
    band_row = None
    for r in range(1, ws.max_row + 1):
        v = _v(ws.cell(row=r, column=1))
        if v.startswith(band_keyword):
            band_row = r
            break
    if band_row is None:
        return None, 0
    header_row = band_row + 1
    count = 0
    r = header_row + 1
    while r <= ws.max_row:
        v_a = _v(ws.cell(row=r, column=1))
        if v_a.startswith("Packing instructions") or v_a.startswith("Dimensions") or v_a.startswith("Artwork files"):
            break
        if not v_a and not _v(ws.cell(row=r, column=2)) and not _v(ws.cell(row=r, column=3)):
            # tolerate ONE blank row
            r2 = r + 1
            if r2 > ws.max_row or not any(_v(ws.cell(row=r2, column=c)) for c in range(1, 6)):
                break
        count += 1
        r += 1
    return header_row, count


def read_component_tab(ws, tab_name: str, display_name: str, style: str) -> Component:
    comp = Component(tab_name=tab_name, display_name=display_name, description="",
                     page_order=9999, style=style)
    comp.description = _v(ws.cell(row=5, column=2))
    comp.pdf_page_title = _v(ws.cell(row=6, column=2)) or display_name

    # Specs — single column. Field labels start at row 10, value in column B.
    for r in range(10, ws.max_row + 1):
        label = _v(ws.cell(row=r, column=1))
        if not label:
            break
        value = _v(ws.cell(row=r, column=2))
        comp.specs[label] = value
        if label.startswith("Notes"):
            break

    # Artwork files
    art_header, art_count = _find_block(ws, "Artwork files")
    if art_header:
        for i in range(1, art_count + 1):
            rr = art_header + i
            atype = _v(ws.cell(row=rr, column=1))
            caption = _v(ws.cell(row=rr, column=2))
            fname = _v(ws.cell(row=rr, column=3))
            if atype or fname or caption:
                comp.artworks.append(ArtworkEntry(atype, caption, fname))

    # Packing instructions
    pi_header, pi_count = _find_block(ws, "Packing instructions")
    if pi_header:
        for i in range(1, pi_count + 1):
            rr = pi_header + i
            step = _v(ws.cell(row=rr, column=1))
            instr = _v(ws.cell(row=rr, column=2))
            fname = _v(ws.cell(row=rr, column=3))
            if instr or fname:
                comp.packing_steps.append(PackingStep(step, instr, fname))

    # Dimensions
    for r in range(1, ws.max_row + 1):
        if _v(ws.cell(row=r, column=1)).startswith("Dimensions"):
            for rr in range(r + 1, r + 8):
                label = _v(ws.cell(row=rr, column=1))
                value = _v(ws.cell(row=rr, column=2))
                if label and value:
                    comp.dimensions[label] = value
            break
    return comp


def load_project(workbook_path: Path) -> Project:
    wb = load_workbook(workbook_path, data_only=True)
    proj = Project()

    if "Project Info" in wb.sheetnames:
        info = read_project_info(wb["Project Info"])
        proj.project_name = info.get("Project Name", "")
        proj.product_type = info.get("Product Type", "")
        proj.product_family = info.get("Product Family", "")
        proj.sku_colourway = info.get("SKU / Colourway", "")
        proj.designer = info.get("Packaging Designer", "")
        proj.engineer = info.get("Packaging Engineer", "")
        proj.graphic_designer = info.get("Graphic Designer", "")
        proj.brand_manager = info.get("Brand Manager", "")
        proj.date = _format_date_eu(info.get("Date", ""))
        proj.stage = info.get("Project Stage", "")
        proj.supplier = info.get("Supplier", "")
        proj.internal_ref = info.get("Internal Reference", "")
        proj.notes = info.get("Notes", "")
        af = info.get("Artwork Folder", "").strip()
        if af:
            p = Path(af)
            if p.is_dir():
                proj.artwork_folder = p
        proj.overview_image_name = info.get("Packaging Overview Image", "")

    styles_by_tab = (read_components_library(wb["Components Library"])
                     if "Components Library" in wb.sheetnames else {})

    setup = read_product_setup(wb["Product Setup"]) if "Product Setup" in wb.sheetnames else []
    active = sorted(
        [s for s in setup if s["include"] and s["tab_name"] in wb.sheetnames],
        key=lambda s: s["order"],
    )
    for entry in active:
        style = styles_by_tab.get(entry["tab_name"], "single_face")
        comp = read_component_tab(wb[entry["tab_name"]], entry["tab_name"],
                                  entry["display"], style)
        comp.page_order = entry["order"]
        proj.components.append(comp)
    return proj


# -- File-name resolution ---------------------------------------------------
def resolve_image(folder: Path | None, name: str, workbook_dir: Path | None = None) -> Path | None:
    """Resolve a file reference (absolute path, relative path, name, stem, or glob)
    against `folder`, searching recursively. Archive/ paths are deprioritised."""
    if not name:
        return None
    q = name.strip().strip('"').strip("'")
    p = Path(q)
    if p.is_absolute() and p.exists():
        return p
    if workbook_dir and (workbook_dir / q).exists():
        return workbook_dir / q
    if not (folder and folder.is_dir()):
        return None
    if (folder / q).exists():
        return folder / q

    def _rank(path: Path) -> tuple[int, int]:
        depth = len(path.relative_to(folder).parts)
        archived = any("archive" in part.lower() for part in path.parts)
        return (1 if archived else 0, depth)

    def _hits(patterns: list[str]) -> list[Path]:
        out = []
        for pat in patterns:
            for f in folder.rglob(pat):
                if f.is_file() and f.suffix.lower() in IMAGE_EXTS:
                    out.append(f)
        return sorted(set(out), key=_rank)

    if any(ch in q for ch in "*?["):
        h = _hits([q]); return h[0] if h else None
    h = _hits([q])
    if h: return h[0]
    h = _hits([f"{q}{ext}" for ext in IMAGE_EXTS])
    if h: return h[0]
    h = _hits([f"{q}*"])
    return h[0] if h else None


# -- PDF helpers ------------------------------------------------------------
def _wrap(c, text, font, size, max_w):
    if not text:
        return [""]
    words = text.split()
    lines, cur = [], ""
    for w in words:
        trial = (cur + " " + w).strip()
        if c.stringWidth(trial, font, size) <= max_w:
            cur = trial
        else:
            if cur:
                lines.append(cur)
            cur = w
    if cur:
        lines.append(cur)
    return lines


def _header_strip(c, project: Project, page_title: str):
    top = PAGE_H - MARGIN
    c.setFillColor(INK)
    c.setFont(FONT_BOLD, 22)
    c.drawString(MARGIN, top - 8, "loop")

    x = MARGIN + 70
    c.setFont(FONT_BOLD, 8); c.setFillColor(INK)
    c.drawString(x, top - 2, "PROJECT NAME:")
    c.setFont(FONT_REG, 8); c.drawString(x + 70, top - 2, project.project_name)
    c.setFont(FONT_BOLD, 8); c.drawString(x, top - 14, "PAGE NAME:")
    c.setFont(FONT_REG, 8); c.drawString(x + 70, top - 14, page_title)
    c.setFont(FONT_BOLD, 8); c.drawString(x, top - 26, "DATE:")
    c.setFont(FONT_REG, 8); c.drawString(x + 70, top - 26, project.date)

    rx = PAGE_W / 2 + 30
    # Right column: Designer, Engineer, (optional) Graphic Designer, SKU.
    # Tighten the step when four rows are present so all stay above the divider.
    rows = [("PACKAGING DESIGNER:", project.designer),
            ("PACKAGING ENGINEER:", project.engineer)]
    if project.graphic_designer:
        rows.append(("GRAPHIC DESIGNER:", project.graphic_designer))
    if project.sku_colourway:
        rows.append(("SKU / COLOURWAY:", project.sku_colourway))
    _step = 10.5 if len(rows) > 3 else 12
    for i, (lbl, val) in enumerate(rows):
        yy = top - 2 - i * _step
        c.setFont(FONT_BOLD, 8); c.drawString(rx, yy, lbl)
        c.setFont(FONT_REG, 8); c.drawString(rx + 110, yy, val)

    c.setFont(FONT_REG, 9); c.setFillColor(INK_MID)
    c.drawRightString(PAGE_W - MARGIN, top - 2, "Packaging Creative Intent")
    if project.stage:
        bw, bh = 60, 18
        bx = PAGE_W - MARGIN - bw
        by = top - 28
        c.setFillColor(INK_DARK); c.rect(bx, by, bw, bh, stroke=0, fill=1)
        c.setFillColor(white); c.setFont(FONT_BOLD, 11)
        c.drawCentredString(bx + bw / 2, by + 5, project.stage)

    c.setStrokeColor(LINE); c.setLineWidth(0.5)
    c.line(MARGIN, top - 36, PAGE_W - MARGIN, top - 36)


def _image_fit(c, src_path: Path | None, x, y, max_w, max_h, *, caption: str = "",
               file_name_display: str = "", placeholder: str = "[no artwork]"):
    if src_path is None or not Path(src_path).exists():
        c.setStrokeColor(LINE); c.setFillColor(BAND)
        c.rect(x, y, max_w, max_h, stroke=1, fill=1)
        c.setFillColor(INK_MID); c.setFont(FONT_ITALIC, 9)
        c.drawCentredString(x + max_w / 2, y + max_h / 2, placeholder)
        text_y = y - 10
        if caption:
            c.setFont(FONT_ITALIC, 8)
            c.drawCentredString(x + max_w / 2, text_y, caption)
            text_y -= 9
        if file_name_display:
            c.setFont(FONT_MONO, 7.5); c.setFillColor(INK_MID)
            c.drawCentredString(x + max_w / 2, text_y, file_name_display)
        return

    src = Path(src_path)
    if src.suffix.lower() == ".pdf":
        if shutil.which("pdftoppm"):
            tmp = Path(tempfile.mkdtemp(prefix="loop_img_"))
            out = tmp / "render"
            subprocess.run(["pdftoppm", "-r", "150", "-png", "-f", "1", "-l", "1",
                            "-singlefile", str(src), str(out)], check=False)
            src = out.with_suffix(".png")
        else:
            return _image_fit(c, None, x, y, max_w, max_h, caption=caption,
                              file_name_display=file_name_display,
                              placeholder="[install pdftoppm]")

    try:
        with PILImage.open(src) as im:
            iw, ih = im.size
    except Exception:
        return _image_fit(c, None, x, y, max_w, max_h, caption=caption,
                          file_name_display=file_name_display,
                          placeholder=f"[cannot open {src.name}]")

    ratio = min(max_w / iw, max_h / ih)
    dw = iw * ratio; dh = ih * ratio
    dx = x + (max_w - dw) / 2
    dy = y + (max_h - dh) / 2
    c.drawImage(str(src), dx, dy, width=dw, height=dh,
                preserveAspectRatio=True, mask="auto")
    text_y = y - 10
    if caption:
        c.setFillColor(INK_MID); c.setFont(FONT_ITALIC, 8)
        c.drawCentredString(x + max_w / 2, text_y, caption)
        text_y -= 9
    if file_name_display:
        c.setFont(FONT_MONO, 7.5); c.setFillColor(INK_MID)
        c.drawCentredString(x + max_w / 2, text_y, file_name_display)


def _draw_field(c, x, y, max_w, label, value, *, label_size=8, value_size=8, line_h=11):
    c.setFillColor(INK); c.setFont(FONT_BOLD, label_size)
    label_text = f"{label}:"
    c.drawString(x, y, label_text)
    label_w = c.stringWidth(label_text, FONT_BOLD, label_size) + 4
    if not value:
        return y - line_h
    c.setFont(FONT_REG, value_size); c.setFillColor(INK_DARK)
    lines = _wrap(c, value, FONT_REG, value_size, max_w - label_w)
    c.drawString(x + label_w, y, lines[0])
    cur_y = y
    for extra in lines[1:]:
        cur_y -= line_h
        c.drawString(x + label_w, cur_y, extra)
    return cur_y - line_h


# -- Page renderers ---------------------------------------------------------
def render_overview(c, project: Project):
    _header_strip(c, project, "Packaging Overview")
    top = PAGE_H - MARGIN - 50

    # Title
    c.setFillColor(INK); c.setFont(FONT_BOLD, 16)
    c.drawString(MARGIN, top, "Packaging Overview")
    c.setFont(FONT_REG, 9); c.setFillColor(INK_MID)
    sku = f" · SKU: {project.sku_colourway}" if project.sku_colourway else ""
    c.drawString(MARGIN, top - 16,
                 f"{len(project.components)} component(s) · "
                 f"{project.product_type or 'Product'}{sku}")

    # Left: component key  |  Right: exploded image
    list_w = 220
    list_x = MARGIN
    list_top = top - 40
    list_bot = MARGIN + 20
    img_x = MARGIN + list_w + 30
    img_w = PAGE_W - MARGIN - img_x
    img_top = list_top
    img_bot = list_bot

    # Component list
    c.setFillColor(INK); c.setFont(FONT_BOLD, 9)
    c.drawString(list_x, list_top, "COMPONENTS")
    y = list_top - 16
    for comp in project.components:
        # Bullet
        c.setFillColor(INK)
        c.circle(list_x + 3, y + 3, 2, stroke=0, fill=1)
        c.setFont(FONT_BOLD, 10)
        c.drawString(list_x + 12, y, comp.display_name)
        # Description
        c.setFillColor(INK_MID); c.setFont(FONT_REG, 8)
        desc_lines = _wrap(c, comp.description, FONT_REG, 8, list_w - 14)
        for j, line in enumerate(desc_lines[:2]):
            c.drawString(list_x + 12, y - 11 - j * 10, line)
        y -= 36

    # Right: exploded image (or placeholder)
    img = resolve_image(project.artwork_folder, project.overview_image_name)
    label = project.overview_image_name or "(set 'Packaging Overview Image' in Project Info)"
    _image_fit(c, img, img_x, img_bot, img_w, img_top - img_bot,
               caption="", file_name_display=label,
               placeholder=f"[Place '{project.overview_image_name}' in your Artwork Folder]"
                           if project.overview_image_name else "[no overview image configured]")


def _find_artworks_by_type(comp: Component, target_types: list[str]) -> list[ArtworkEntry]:
    """Return the artwork rows matching a target type (case-insensitive, fuzzy)."""
    out = []
    for tt in target_types:
        for a in comp.artworks:
            if tt.lower().replace("_", " ") in a.artwork_type.lower().replace("_", " "):
                if a.file_name or a.caption:
                    out.append(a)
                    break
    return out


def render_component_two_face(c, project: Project, comp: Component):
    """Page 2-3 style: single-option spec + Mockup + Artwork_Front + (optional) Artwork_Back."""
    _header_strip(c, project, comp.pdf_page_title or comp.display_name)
    top = PAGE_H - MARGIN - 50

    # Spec block (full width)
    c.setFillColor(ACCENT); c.setFont(FONT_BOLD, 11)
    c.drawString(MARGIN, top, "SPECIFICATIONS")
    y = top - 16
    spec_w = PAGE_W - 2 * MARGIN
    for field_label, display_label in RENDERED_SPEC_LINES:
        v = comp.specs.get(field_label, "")
        if not v:
            continue
        y = _draw_field(c, MARGIN, y, spec_w, display_label, v,
                        label_size=9, value_size=9, line_h=12)
    note = comp.specs.get("Notes", "")
    if note:
        y -= 4
        c.setFillColor(INK_MID); c.setFont(FONT_ITALIC, 8)
        for ln in _wrap(c, f"Note: {note}", FONT_ITALIC, 8, spec_w)[:3]:
            c.drawString(MARGIN, y, ln); y -= 11

    # Divider
    div_y = y - 6
    c.setStrokeColor(LINE); c.setLineWidth(0.5)
    c.line(MARGIN, div_y, PAGE_W - MARGIN, div_y)

    # Find the three slots
    mock = next((a for a in comp.artworks if "mockup" in a.artwork_type.lower().replace("_", " ").replace("-", " ")), None)
    front = next((a for a in comp.artworks if "front" in a.artwork_type.lower()), None)
    back = next((a for a in comp.artworks if "back" in a.artwork_type.lower()), None)

    slots = [("MOCKUP", mock), ("ARTWORK – FRONT", front)]
    if back and (back.file_name or back.caption):
        slots.append(("ARTWORK – BACK", back))

    # Section title
    c.setFillColor(INK); c.setFont(FONT_BOLD, 10)
    c.drawString(MARGIN, div_y - 16, "ARTWORK & MOCKUP")

    grid_top = div_y - 24
    grid_bot = MARGIN + 20
    grid_h = grid_top - grid_bot
    n = len(slots)
    cell_w = (PAGE_W - 2 * MARGIN - 16 * (n - 1)) / n
    cell_h_max = grid_h - 40  # leave room for label and filename

    for i, (title, art) in enumerate(slots):
        cx = MARGIN + i * (cell_w + 16)
        label_y = grid_top - 4
        cell_top = label_y - 6
        cell_h = cell_h_max
        cell_bot = cell_top - cell_h
        # Title
        c.setFillColor(INK); c.setFont(FONT_BOLD, 8.5)
        c.drawString(cx, label_y, title)
        # Image
        img = resolve_image(project.artwork_folder, art.file_name) if art else None
        _image_fit(c, img, cx, cell_bot + 24, cell_w, cell_h - 24,
                   caption=art.caption if art else "",
                   file_name_display=art.file_name if art else "(no file specified)")


def render_component_single_face(c, project: Project, comp: Component):
    """Page 4-6 style: single-option spec + Mockup + Artwork + Packing Instructions."""
    _header_strip(c, project, comp.pdf_page_title or comp.display_name)
    top = PAGE_H - MARGIN - 50

    # Spec block (left column ~55%); right column reserved for packing or extra
    spec_w_total = PAGE_W - 2 * MARGIN

    c.setFillColor(ACCENT); c.setFont(FONT_BOLD, 11)
    c.drawString(MARGIN, top, "SPECIFICATIONS")
    y = top - 16
    for field_label, display_label in RENDERED_SPEC_LINES:
        v = comp.specs.get(field_label, "")
        if not v:
            continue
        y = _draw_field(c, MARGIN, y, spec_w_total, display_label, v,
                        label_size=9, value_size=9, line_h=12)
    note = comp.specs.get("Notes", "")
    if note:
        y -= 4
        c.setFillColor(INK_MID); c.setFont(FONT_ITALIC, 8)
        for ln in _wrap(c, f"Note: {note}", FONT_ITALIC, 8, spec_w_total)[:3]:
            c.drawString(MARGIN, y, ln); y -= 11

    div_y = y - 6
    c.setStrokeColor(LINE); c.setLineWidth(0.5)
    c.line(MARGIN, div_y, PAGE_W - MARGIN, div_y)

    # Lower zone: split into LEFT (artwork) and RIGHT (packing instructions)
    has_packing = bool(comp.packing_steps)
    if has_packing:
        left_w = (PAGE_W - 2 * MARGIN) * 0.55 - 10
    else:
        left_w = PAGE_W - 2 * MARGIN
    left_x = MARGIN
    right_x = MARGIN + left_w + 20
    right_w = PAGE_W - MARGIN - right_x

    grid_top = div_y - 18
    grid_bot = MARGIN + 20

    # ---- Left: Mockup + Artwork side by side
    c.setFillColor(INK); c.setFont(FONT_BOLD, 10)
    c.drawString(left_x, grid_top + 4, "ARTWORK & MOCKUP")

    mock = next((a for a in comp.artworks if "mockup" in a.artwork_type.lower().replace("_", " ").replace("-", " ")), None)
    artw = next((a for a in comp.artworks if a is not mock and "artwork" in a.artwork_type.lower()), None)
    if artw is None:
        # any remaining non-mockup row
        artw = next((a for a in comp.artworks
                     if (a.file_name or a.caption) and a is not mock), None)

    slots = []
    if mock:
        slots.append(("MOCKUP", mock))
    if artw:
        slots.append(("ARTWORK", artw))
    if not slots:
        slots = [("MOCKUP", None), ("ARTWORK", None)]

    n = max(len(slots), 1)
    cell_w = (left_w - 16 * (n - 1)) / n
    cell_h_total = grid_top - grid_bot
    cell_h = max(70, cell_h_total - 50)

    for i, (title, art) in enumerate(slots):
        cx = left_x + i * (cell_w + 16)
        label_y = grid_top - 8
        cell_top = label_y - 6
        cell_bot = cell_top - cell_h
        c.setFillColor(INK); c.setFont(FONT_BOLD, 8.5)
        c.drawString(cx, label_y, title)
        img = resolve_image(project.artwork_folder, art.file_name) if art else None
        _image_fit(c, img, cx, cell_bot + 24, cell_w, cell_h - 24,
                   caption=art.caption if art else "",
                   file_name_display=(art.file_name if art and art.file_name
                                      else "(no file specified)"))

    # ---- Right: Packing instructions (text + optional image per step)
    if has_packing:
        c.setFillColor(INK); c.setFont(FONT_BOLD, 10)
        c.drawString(right_x, grid_top + 4, "PACKING INSTRUCTIONS")

        # Stack steps top-down. Each step gets a small thumbnail (if present) +
        # text wrapped to the column.
        y_cursor = grid_top - 12
        thumb_w = 70
        thumb_h = 50
        gap = 6
        for step in comp.packing_steps:
            if y_cursor - thumb_h - 6 < grid_bot:
                break
            # Step number / label
            c.setFillColor(ACCENT); c.setFont(FONT_BOLD, 8.5)
            c.drawString(right_x, y_cursor, (step.step or "").upper())
            # Thumb (if any)
            img = resolve_image(project.artwork_folder, step.file_name)
            if img is not None:
                _image_fit(c, img, right_x, y_cursor - 6 - thumb_h, thumb_w, thumb_h,
                           caption="", file_name_display="")
                text_x = right_x + thumb_w + gap
                text_w = right_w - thumb_w - gap
            else:
                text_x = right_x
                text_w = right_w
            # Instruction text
            c.setFillColor(INK_DARK); c.setFont(FONT_REG, 8.5)
            instr_y = y_cursor - 10
            lines = _wrap(c, step.instruction, FONT_REG, 8.5, text_w)
            for ln in lines[:5]:
                c.drawString(text_x, instr_y, ln); instr_y -= 11
            # File name as supplier-search reference
            if step.file_name:
                c.setFont(FONT_MONO, 7); c.setFillColor(INK_MID)
                c.drawString(text_x, instr_y, step.file_name)
                instr_y -= 10
            # Bottom of this step block — whichever is lower of thumb-bottom or text-end
            block_bot = min((y_cursor - thumb_h - 8) if img else y_cursor,
                            instr_y)
            y_cursor = block_bot - 8


# -- Main -------------------------------------------------------------------
def generate(workbook_path: Path, output_path: Path):
    project = load_project(workbook_path)
    c = canvas.Canvas(str(output_path), pagesize=(PAGE_W, PAGE_H))
    title_parts = [project.project_name or "Packaging"]
    if project.sku_colourway:
        title_parts.append(project.sku_colourway)
    c.setTitle(" — ".join(title_parts) + " — Packaging Creative Intent")
    c.setAuthor(project.designer or "Loop")
    c.setSubject("Packaging Creative Intent")

    render_overview(c, project)
    c.showPage()

    for comp in project.components:
        if comp.style == "two_face":
            render_component_two_face(c, project, comp)
        else:
            render_component_single_face(c, project, comp)
        c.showPage()

    c.save()
    print(f"Saved PDF: {output_path}")


def main(argv=None):
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("workbook", type=Path)
    p.add_argument("output", type=Path, nargs="?", default=None)
    args = p.parse_args(argv)
    out = args.output or args.workbook.with_suffix("").with_name(
        args.workbook.stem.replace("_EXAMPLE", "").replace("_TEMPLATE", "") + "_OUTPUT.pdf"
    )
    generate(args.workbook, out)


if __name__ == "__main__":
    main()
