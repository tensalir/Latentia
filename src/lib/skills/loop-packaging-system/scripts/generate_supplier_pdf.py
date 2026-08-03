"""
Generate a supplier-ready PDF from an editable Illustrator file + workbook data.

Three modes (one PDF each) so Ana can compare:

  Option A  Info-box OVERLAID onto the artwork page, in the white margin.
            Original artwork stream is left untouched; the info-box content is
            stamped on a new content layer above the existing page.
            Output: same page count as the source artwork.

  Option B  Artwork fonts OUTLINED via Ghostscript, then the info-box page
            APPENDED as an extra page. Production-grade if your printer
            requires fully outlined fonts.
            Output: source pages + 1.

  Option C  Artwork left COMPLETELY UNTOUCHED (fonts stay embedded, layers and
            spot-color plates intact), info-box page APPENDED at the end.
            Output: source pages + 1.

In every mode, the info box pulls:
  - Material, Print Method, Drawing/Print Part #, Project metadata  →  Excel
  - Ink list, Special Effects/Finishes                              →  AI file PlateNames metadata
"""

from __future__ import annotations

import argparse
import io
import shutil
import subprocess
import tempfile
from pathlib import Path

import pikepdf
from openpyxl import load_workbook
from reportlab.lib.colors import HexColor, white, black
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas

# ---------------------------------------------------------------------------
# Style
# ---------------------------------------------------------------------------
INK = HexColor("#1A1A1A")
INK_MID = HexColor("#6B6B6B")
LINE = HexColor("#D0D0D0")
ACCENT = HexColor("#C8102E")
BAND = HexColor("#F2F2F2")

FONT_BOLD = "Helvetica-Bold"
FONT_REG = "Helvetica"
FONT_MONO = "Courier"
FONT_ITALIC = "Helvetica-Oblique"

# Fields we extract from PlateNames.
DIELINE_KEYWORDS = ("CUT LINE", "BEND LINE", "DIELINE", "DIE CUT", "DIE-CUT",
                    "PERF", "FOLD LINE", "CREASE", "GLUE AREA", "GLUE ZONE", "GLUE")
FINISH_KEYWORDS = ("EMBOSS", "DEBOSS", "UV", "FOIL", "SPOT", "VARNISH", "LAMINATE", "GLOSS", "MATT", "MATTE")


# ---------------------------------------------------------------------------
# Read AI file metadata (no modification)
# ---------------------------------------------------------------------------
def extract_plate_info(ai_path: Path) -> dict:
    """Return {'inks': [...], 'finishes': [...], 'dielines': [...]} from PlateNames."""
    with pikepdf.open(ai_path) as pdf:
        with pdf.open_metadata() as md:
            plates = md.get("{http://ns.adobe.com/xap/1.0/t/pg/}PlateNames", []) or []
    inks, finishes, dielines = [], [], []
    for p in plates:
        up = str(p).strip().upper()
        if any(k in up for k in DIELINE_KEYWORDS):
            dielines.append(str(p))
        elif any(k in up for k in FINISH_KEYWORDS):
            finishes.append(str(p))
        else:
            inks.append(str(p))
    return {"inks": inks, "finishes": finishes, "dielines": dielines}


# ---------------------------------------------------------------------------
# Read Excel
# ---------------------------------------------------------------------------
def _v(c):
    return "" if c.value is None else str(c.value).strip()


def _format_date_eu(raw) -> str:
    """DD-MM-YYYY, no time. Accepts datetime/date/string; falls back gracefully."""
    import datetime as _dt
    if raw is None or raw == "":
        return ""
    if isinstance(raw, (_dt.datetime, _dt.date)):
        return raw.strftime("%d-%m-%Y")
    s = str(raw).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y",
                "%m/%d/%Y", "%Y/%m/%d", "%d.%m.%Y"):
        try:
            return _dt.datetime.strptime(s, fmt).strftime("%d-%m-%Y")
        except ValueError:
            continue
    return s.split(" ")[0]


def read_workbook_for_component(workbook_path: Path, component_tab: str) -> dict:
    wb = load_workbook(workbook_path, data_only=True)
    pi = wb["Project Info"] if "Project Info" in wb.sheetnames else None
    info = {}
    if pi:
        for r in range(5, pi.max_row + 1):
            k = _v(pi.cell(row=r, column=2))
            v = _v(pi.cell(row=r, column=3))
            if k:
                info[k] = v
    if "Date" in info:
        info["Date"] = _format_date_eu(info["Date"])

    comp = {}
    if component_tab in wb.sheetnames:
        ws = wb[component_tab]
        comp["display_name"] = _v(ws.cell(row=4, column=2)) or component_tab
        for r in range(10, ws.max_row + 1):
            label = _v(ws.cell(row=r, column=1))
            if not label:
                break
            comp[label] = _v(ws.cell(row=r, column=2))
            if label == "Notes":
                break
    return {"project": info, "component": comp}


# ---------------------------------------------------------------------------
# Build the info-box content as a PDF (single page or stamp)
# ---------------------------------------------------------------------------
def render_info_box_full_page(out_path: Path, *, project: dict, component: dict,
                              plate_info: dict, component_display: str,
                              page_size=None):
    """Create a one-page PDF that mirrors the Nyx Creative Intent layout."""
    if page_size is None:
        page_size = landscape(A4)
    PAGE_W, PAGE_H = page_size
    M = 16 * mm
    c = canvas.Canvas(str(out_path), pagesize=page_size)

    # Header strip
    top = PAGE_H - M
    c.setFillColor(INK); c.setFont(FONT_BOLD, 22)
    c.drawString(M, top - 8, "loop")

    def kv(x, y, label, value, *, label_w=110):
        c.setFont(FONT_BOLD, 8); c.setFillColor(INK)
        c.drawString(x, y, label)
        c.setFont(FONT_REG, 8); c.setFillColor(INK)
        c.drawString(x + label_w, y, value or "")

    x1 = M + 70
    kv(x1, top - 2, "PROJECT NAME:", project.get("Project Name", ""))
    kv(x1, top - 14, "PART NAME:", component_display)
    kv(x1, top - 26, "DATE:", project.get("Date", ""))

    x2 = PAGE_W / 2 + 30
    kv(x2, top - 2,  "DRAWING PART No.:", component.get("Drawing Part Number", ""))
    kv(x2, top - 14, "PRINT PART No.:",   component.get("Print Part Number", ""))
    kv(x2, top - 26, "SKU / COLOURWAY:",  project.get("SKU / Colourway", ""))

    c.setFont(FONT_REG, 9); c.setFillColor(INK_MID)
    c.drawRightString(PAGE_W - M, top - 2, "Printing Supplier Brief")
    if project.get("Project Stage"):
        bw, bh = 60, 18
        bx, by = PAGE_W - M - bw, top - 28
        c.setFillColor(HexColor("#262626")); c.rect(bx, by, bw, bh, stroke=0, fill=1)
        c.setFillColor(white); c.setFont(FONT_BOLD, 11)
        c.drawCentredString(bx + bw / 2, by + 5, project["Project Stage"])

    c.setStrokeColor(LINE); c.setLineWidth(0.5)
    c.line(M, top - 36, PAGE_W - M, top - 36)

    # Title
    y = top - 56
    c.setFillColor(ACCENT); c.setFont(FONT_BOLD, 13)
    c.drawString(M, y, f"{component_display} — Printing Brief")

    # Two columns
    col_w = (PAGE_W - 2 * M) / 2 - 10
    left_x = M
    right_x = M + col_w + 20

    # LEFT — Spec (from Excel)
    y = y - 30
    c.setFillColor(INK); c.setFont(FONT_BOLD, 10)
    c.drawString(left_x, y, "MATERIAL & PROCESS")
    y -= 16
    def spec_row(label, value):
        nonlocal y
        c.setFillColor(INK); c.setFont(FONT_BOLD, 9)
        c.drawString(left_x, y, label)
        c.setFont(FONT_REG, 9); c.setFillColor(HexColor("#333333"))
        # wrap value
        max_w = col_w - 130
        text = value or "—"
        words = text.split()
        line, lines = "", []
        for w in words:
            trial = (line + " " + w).strip()
            if c.stringWidth(trial, FONT_REG, 9) <= max_w:
                line = trial
            else:
                if line:
                    lines.append(line)
                line = w
        if line:
            lines.append(line)
        for i, ln in enumerate(lines[:4]):
            c.drawString(left_x + 130, y - i * 11, ln)
        y -= max(14, 11 * min(len(lines), 4) + 4)

    spec_row("Material:",       component.get("Material", ""))
    spec_row("Printing Method:", component.get("Printing Method", ""))
    spec_row("Inks (specified):", component.get("Inks / Print", ""))
    spec_row("Finishes (specified):", component.get("Finishes", ""))
    spec_row("MSDS reference:", component.get("Coating MSDS Ref.", ""))

    # RIGHT — Inks & Finishes (extracted from AI file PlateNames)
    y_right = top - 86
    c.setFillColor(INK); c.setFont(FONT_BOLD, 10)
    c.drawString(right_x, y_right, "INKS & FINISHES (auto-detected from AI file)")
    y_right -= 16

    def chips(title, items, color):
        nonlocal y_right
        c.setFillColor(INK); c.setFont(FONT_BOLD, 9)
        c.drawString(right_x, y_right, title)
        y_right -= 14
        cx = right_x
        cy = y_right
        for item in items:
            tw = c.stringWidth(item, FONT_REG, 8.5)
            chip_w = tw + 16
            if cx + chip_w > PAGE_W - M:
                cx = right_x; cy -= 24
            c.setFillColor(color); c.setStrokeColor(color)
            c.roundRect(cx, cy - 4, chip_w, 18, 4, stroke=0, fill=1)
            c.setFillColor(white); c.setFont(FONT_REG, 8.5)
            c.drawString(cx + 8, cy + 2, item)
            cx += chip_w + 6
        y_right = cy - 18

    if plate_info["inks"]:
        chips(f"Inks ({len(plate_info['inks'])})", plate_info["inks"], INK)
    else:
        c.setFillColor(INK_MID); c.setFont(FONT_ITALIC, 9)
        c.drawString(right_x, y_right, "No inks detected in PlateNames metadata.")
        y_right -= 16

    if plate_info["finishes"]:
        chips(f"Special finishes ({len(plate_info['finishes'])})", plate_info["finishes"], ACCENT)
    if plate_info["dielines"]:
        chips(f"Structural plates ({len(plate_info['dielines'])})", plate_info["dielines"], HexColor("#6B6B6B"))

    # Approval & footer
    foot_y = M + 24
    c.setStrokeColor(LINE); c.setLineWidth(0.5)
    c.line(M, foot_y + 16, PAGE_W - M, foot_y + 16)
    c.setFillColor(INK_MID); c.setFont(FONT_REG, 8)
    c.drawString(M, foot_y, f"Designer: {project.get('Packaging Designer','')} · "
                            f"Engineer: {project.get('Packaging Engineer','')} · "
                            f"Approval: {component.get('Approval Status','')}")
    c.drawRightString(PAGE_W - M, foot_y, "Brief generated from Loop Packaging Creative Intent template")

    c.save()


def render_info_overlay_stamp(out_path: Path, page_size, *, project, component,
                              plate_info, component_display):
    """Render the info-box overlay sized to fit the top strip of the artwork
    page. Designed for Option A: overlay onto the existing artwork.

    Layout — full-width header strip:
      ┌──────────────────────────────────────────────────────────────────┐
      │ loop   PROJECT / PART / DATE      DESIGNER / ENGINEER     [DVT]  │
      │ ────────────────────────────────────────────────────────────────  │
      │ Spec fields (2-col)            │ INKS  [chips]                    │
      │                                │ FINISHES  [chips]                │
      │                                │ DIELINES  [chips]                │
      └──────────────────────────────────────────────────────────────────┘
    """
    PAGE_W, PAGE_H = page_size
    c = canvas.Canvas(str(out_path), pagesize=(PAGE_W, PAGE_H))

    # Fixed physical dimensions — same on every artwork sheet regardless of size.
    # 200 mm wide × 100 mm tall, positioned top-RIGHT with a 10 mm margin from
    # the page edges. Sized to fit on A4 portrait (Ana's smallest sheet).
    MM = mm  # imported above; 1 mm = 2.8346 pt
    BOX_W_MM = 200
    BOX_H_MM = 100
    MARGIN_MM = 10
    box_w = BOX_W_MM * MM
    box_h = BOX_H_MM * MM
    bx = PAGE_W - box_w - MARGIN_MM * MM
    by = PAGE_H - box_h - MARGIN_MM * MM

    # Background card
    c.setFillColor(HexColor("#FFFFFF"))
    c.setStrokeColor(HexColor("#1A1A1A"))
    c.setLineWidth(0.8)
    c.rect(bx, by, box_w, box_h, stroke=1, fill=1)

    pad_x = 16
    pad_y = 14
    inner_left  = bx + pad_x
    inner_right = bx + box_w - pad_x
    top_y       = by + box_h - pad_y

    # --- Header strip inside the box --------------------------------------
    c.setFillColor(INK); c.setFont(FONT_BOLD, 22)
    c.drawString(inner_left, top_y - 16, "loop")

    def kv(x, y, label, value, *, label_w=70, size=8, max_val_w=None):
        c.setFont(FONT_BOLD, size); c.setFillColor(INK)
        c.drawString(x, y, label)
        c.setFont(FONT_REG, size); c.setFillColor(HexColor("#262626"))
        val = value or ""
        if max_val_w:
            while val and c.stringWidth(val, FONT_REG, size) > max_val_w:
                val = val[:-1]
            if val != (value or ""):
                val = val.rstrip() + "…"
        c.drawString(x + label_w, y, val)

    h_y = top_y - 4
    inner_w = box_w - 2 * pad_x
    x_col_a = inner_left + 62
    x_col_b = inner_left + inner_w * 0.45
    RLBL = 106  # width for the longer right-column labels
    # value width budgets so columns never collide (keep clear of the stage badge)
    col_a_val_w = (x_col_b - x_col_a) - 70 - 8
    col_b_val_w = (inner_right - 56 - 8) - (x_col_b + RLBL)

    kv(x_col_a, h_y - 2,  "PROJECT NAME:", project.get("Project Name", ""), max_val_w=col_a_val_w)
    kv(x_col_a, h_y - 16, "PART NAME:",    component_display, max_val_w=col_a_val_w)
    kv(x_col_a, h_y - 30, "DATE:",         project.get("Date", ""), max_val_w=col_a_val_w)

    kv(x_col_b, h_y - 2,  "PACKAGING DESIGNER:", project.get("Packaging Designer", ""), label_w=RLBL, max_val_w=col_b_val_w)
    kv(x_col_b, h_y - 16, "PACKAGING ENGINEER:", project.get("Packaging Engineer", ""), label_w=RLBL, max_val_w=col_b_val_w)
    kv(x_col_b, h_y - 30, "GRAPHIC DESIGNER:",   project.get("Graphic Designer", ""), label_w=RLBL, max_val_w=col_b_val_w)

    # Stage badge top-right
    if project.get("Project Stage"):
        bw, bh = 56, 20
        bxr = inner_right - bw
        byr = top_y - bh
        c.setFillColor(HexColor("#262626"))
        c.rect(bxr, byr, bw, bh, stroke=0, fill=1)
        c.setFillColor(white); c.setFont(FONT_BOLD, 12)
        c.drawCentredString(bxr + bw / 2, byr + 6, project["Project Stage"])

    # Right-edge label (below the designer rows so it never overlaps them)
    c.setFillColor(INK_MID); c.setFont(FONT_REG, 8)
    c.drawRightString(inner_right, top_y - 44, "Printing Brief — auto-generated")

    # Separator
    sep_y = top_y - 48
    c.setStrokeColor(LINE); c.setLineWidth(0.5)
    c.line(inner_left, sep_y, inner_right, sep_y)

    # --- Body — two columns -----------------------------------------------
    body_top = sep_y - 12
    col_gap = 24
    left_col_w  = (box_w - 2 * pad_x - col_gap) * 0.52
    right_col_w = (box_w - 2 * pad_x - col_gap) * 0.48
    right_col_x = inner_left + left_col_w + col_gap

    # LEFT — spec rows
    c.setFillColor(ACCENT); c.setFont(FONT_BOLD, 10)
    c.drawString(inner_left, body_top, "MATERIAL & PROCESS")
    y = body_top - 16

    def spec_row(label, value):
        nonlocal y
        c.setFillColor(INK); c.setFont(FONT_BOLD, 9)
        c.drawString(inner_left, y, label)
        c.setFont(FONT_REG, 9); c.setFillColor(HexColor("#262626"))
        max_w = left_col_w - 110
        text = (value or "—")
        # word-wrap
        words = text.split()
        line, lines = "", []
        for w in words:
            trial = (line + " " + w).strip()
            if c.stringWidth(trial, FONT_REG, 9) <= max_w:
                line = trial
            else:
                if line:
                    lines.append(line)
                line = w
        if line:
            lines.append(line)
        for i, ln in enumerate(lines[:3]):
            c.drawString(inner_left + 110, y - i * 11, ln)
        y -= max(16, 11 * min(len(lines), 3) + 6)

    spec_row("MATERIAL:",   component.get("Material", ""))
    spec_row("METHOD:",     component.get("Printing Method", ""))
    spec_row("MSDS:",       component.get("Coating MSDS Ref.", ""))
    spec_row("SKU CODE:",   project.get("SKU / Colourway", ""))

    # RIGHT — inks / finishes chips
    c.setFillColor(ACCENT); c.setFont(FONT_BOLD, 10)
    c.drawString(right_col_x, body_top, "INKS & FINISHES  (read from AI file)")
    yr = body_top - 16

    def chip_row(title, items, color):
        nonlocal yr
        if not items:
            return
        c.setFillColor(INK); c.setFont(FONT_BOLD, 9)
        c.drawString(right_col_x, yr, title)
        yr -= 20                       # gap between title and first chip
        cx = right_col_x
        cy = yr
        for item in items:
            tw = c.stringWidth(item, FONT_REG, 8.5)
            chip_w = tw + 14
            if cx + chip_w > right_col_x + right_col_w:
                cx = right_col_x; cy -= 26   # gap when chips wrap to a new line
            c.setFillColor(color); c.setStrokeColor(color)
            c.roundRect(cx, cy - 4, chip_w, 16, 4, stroke=0, fill=1)
            c.setFillColor(white); c.setFont(FONT_REG, 8.5)
            c.drawString(cx + 7, cy + 1, item)
            cx += chip_w + 6
        yr = cy - 22                  # gap between this group and the next

    chip_row(f"Inks ({len(plate_info['inks'])})", plate_info["inks"], INK)
    chip_row(f"Special finishes ({len(plate_info['finishes'])})", plate_info["finishes"], ACCENT)
    chip_row(f"Structural plates ({len(plate_info['dielines'])})", plate_info["dielines"], HexColor("#6B6B6B"))

    c.save()


# ---------------------------------------------------------------------------
# Mode A — overlay info box onto artwork page
# ---------------------------------------------------------------------------
def make_option_a(ai_path: Path, out_path: Path, *, project, component, plate_info,
                  component_display):
    """Overlay the info box onto EVERY page of the artwork without modifying its
    content streams. Each page is sized from its own MediaBox, so mixed page
    sizes still get a correctly placed 200x100 mm box top-right."""
    with pikepdf.open(ai_path) as src:
        tmpdir = Path(tempfile.mkdtemp(prefix="loop_ov_"))
        # Cache overlays by (w, h) so identically sized pages reuse one stamp.
        overlay_cache = {}
        for i, page in enumerate(src.pages):
            mb = page.MediaBox
            w = float(mb[2]) - float(mb[0])
            h = float(mb[3]) - float(mb[1])
            key = (round(w, 1), round(h, 1))
            if key not in overlay_cache:
                op = tmpdir / f"_ov_{i}.pdf"
                render_info_overlay_stamp(op, (w, h),
                                          project=project, component=component,
                                          plate_info=plate_info,
                                          component_display=component_display)
                overlay_cache[key] = pikepdf.open(op)
            page.add_overlay(overlay_cache[key].pages[0])
        src.save(out_path)
        for ov in overlay_cache.values():
            ov.close()
    shutil.rmtree(tmpdir, ignore_errors=True)


# ---------------------------------------------------------------------------
# Mode B — outline fonts then append info-box page
# ---------------------------------------------------------------------------
def make_option_b(ai_path: Path, out_path: Path, *, project, component, plate_info,
                  component_display):
    """Ghostscript outline + append info-box page."""
    outlined_path = Path(tempfile.gettempdir()) / "_outlined.pdf"
    subprocess.run([
        "gs", "-o", str(outlined_path), "-sDEVICE=pdfwrite",
        "-dNoOutputFonts",
        "-dColorConversionStrategy=/LeaveColorUnchanged",
        "-dEncodeColorImages=false",
        "-dPDFSETTINGS=/prepress",
        "-dPreserveAnnots=true",
        "-dPreserveOverprintSettings=true",
        "-dCompatibilityLevel=1.6",
        str(ai_path),
    ], check=True, capture_output=True)
    info_path = Path(tempfile.gettempdir()) / "_infoB.pdf"
    with pikepdf.open(outlined_path) as src:
        mb = src.pages[0].MediaBox
        w = float(mb[2]) - float(mb[0]); h = float(mb[3]) - float(mb[1])
        render_info_box_full_page(info_path, project=project, component=component,
                                  plate_info=plate_info, component_display=component_display,
                                  page_size=(w, h))
        with pikepdf.open(info_path) as info:
            src.pages.append(info.pages[0])
        src.save(out_path)
    outlined_path.unlink(missing_ok=True)
    info_path.unlink(missing_ok=True)


# ---------------------------------------------------------------------------
# Mode C — keep artwork untouched, append info-box page
# ---------------------------------------------------------------------------
def make_option_c(ai_path: Path, out_path: Path, *, project, component, plate_info,
                  component_display):
    with pikepdf.open(ai_path) as src:
        mb = src.pages[0].MediaBox
        w = float(mb[2]) - float(mb[0]); h = float(mb[3]) - float(mb[1])
        info_path = Path(tempfile.gettempdir()) / "_infoC.pdf"
        render_info_box_full_page(info_path, project=project, component=component,
                                  plate_info=plate_info, component_display=component_display,
                                  page_size=(w, h))
        with pikepdf.open(info_path) as info:
            src.pages.append(info.pages[0])
        src.save(out_path)
    info_path.unlink(missing_ok=True)


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------
def main():
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--workbook", type=Path, required=True)
    p.add_argument("--ai-file", type=Path, required=True)
    p.add_argument("--component-tab", type=str, default="Outer_Sleeve")
    p.add_argument("--out-dir", type=Path, required=True)
    args = p.parse_args()

    data = read_workbook_for_component(args.workbook, args.component_tab)
    plate_info = extract_plate_info(args.ai_file)
    comp_display = data["component"].get("display_name") or args.component_tab

    print(f"Plate info extracted from {args.ai_file.name}:")
    print(f"  Inks       : {plate_info['inks']}")
    print(f"  Finishes   : {plate_info['finishes']}")
    print(f"  Dielines   : {plate_info['dielines']}")

    args.out_dir.mkdir(parents=True, exist_ok=True)
    stem = args.ai_file.stem.replace(" ", "_")

    out_a = args.out_dir / f"{stem}_OPTION_A_overlay.pdf"

    make_option_a(args.ai_file, out_a, project=data["project"], component=data["component"],
                  plate_info=plate_info, component_display=comp_display)
    print(f"Saved: {out_a}")


if __name__ == "__main__":
    main()
